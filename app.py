from flask import Flask, render_template,redirect,request
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

#importando modulos
from limpieza.limpiar import limpiar_dataset
from analisis.metricas import analizar_kpis, crear_visual, division_categoria, movimientos_fecha
from analisis.insights import insights, alerta_gastos, balance

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    # Obtener el archivo cargado
    archivo = request.files['file']
    
    # Guardar el archivo en la carpeta de uploads
    archivo.save('uploads/' + archivo.filename)

    try:
        # Leer el archivo
        if archivo.filename.endswith('.csv'):
            df_crudo = pd.read_csv('uploads/' + archivo.filename)
        elif archivo.filename.endswith('.xlsx') or archivo.filename.endswith('.xlsm') or archivo.filename.endswith('.xls'):
            df_crudo = pd.read_excel('uploads/' + archivo.filename) 
        else:
            error = "El archivo debe ser .csv, .xlsx, .xls o .xlsm"
            return render_template('/index.html', error=error)

        # Limpieza
        df_limpio = limpiar_dataset(df_crudo)

        # Crear df visual para el frontend
        df_visual = crear_visual(df_limpio)

        # Guardar el archivo limpio en la carpeta de uploads
        df_limpio.to_excel('uploads/limpio.xlsx', index=False)

        # Calcular los KPIs
        kpis = analizar_kpis(df_limpio)
        
        # Calcular las divisiones x categorias
        gastos_x_categoria, ingresos_x_categoria = division_categoria(df_limpio)

        # Calcular los movimientos x fecha
        df_fecha_gasto, df_fecha_ingreso = movimientos_fecha(df_limpio)

        # Calcular la categoria con mas gasto y mas ingreso
        max_gasto, max_ingreso = insights(df_limpio)

        # Calcular la alerta de gastos
        porcentaje = alerta_gastos(df_limpio)

        # Calcular el balance
        balance_mensual = balance(df_limpio)

        return render_template('index.html', kpis=kpis, df_visual=df_visual,gastos_x_categoria=gastos_x_categoria,ingresos_x_categoria=ingresos_x_categoria,df_fecha_gasto=df_fecha_gasto,df_fecha_ingreso=df_fecha_ingreso,max_gasto=max_gasto,max_ingreso=max_ingreso,porcentaje=porcentaje, balance_mensual=balance_mensual)
    
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        return redirect('/')

@app.route('/exportar', methods=['POST'])
def exportar():
    # Lógica para exportar el archivo
    # Leer archivo de uploads
    df_limpio = pd.read_excel('uploads/limpio.xlsx')
    
    # Calcular los KPIs
    kpis = analizar_kpis(df_limpio)

    # Crear reporte financiero
    wb = openpyxl.Workbook()

    # Hoja resumen
    hoja_resumen = wb.active 
    # Título
    hoja_resumen.merge_cells('A1:B1')
    hoja_resumen['A1'] = "REPORTE FINANCIERO"

    hoja_resumen['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    hoja_resumen['A1'].fill = PatternFill("solid", fgColor="1F4E78")
    hoja_resumen['A1'].alignment = Alignment(horizontal="center")

    # Fecha y subtítulos
    hoja_resumen['A4'] = f"Fecha de creación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    hoja_resumen['A4'].font = Font(bold=True)
    hoja_resumen['A6'].font = Font(size=14, bold=True)

    # Encabezados de tabla
    hoja_resumen['A8'] = "Métrica"
    hoja_resumen['B8'] = "Valor"

    encabezado_fill = PatternFill("solid", fgColor="1F4E78")
    encabezado_font = Font(color="FFFFFF", bold=True)

    for celda in ['A8', 'B8']:
        hoja_resumen[celda].fill = encabezado_fill
        hoja_resumen[celda].font = encabezado_font
        hoja_resumen[celda].alignment = Alignment(horizontal="center")

    fila = 9

    # Datos
    for key, value in kpis.items():
        if key.endswith('_raw'):
            continue
        hoja_resumen[f"A{fila}"] = key.replace('_', ' ').title()
        hoja_resumen[f"B{fila}"] = value
        fila += 1

    # Bordes
    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in range(8, fila):
        for col in ['A', 'B']:
            hoja_resumen[f"{col}{row}"].border = borde

    # Alineación
    for row in range(9, fila):
        hoja_resumen[f"A{row}"].alignment = Alignment(horizontal="left")
        hoja_resumen[f"B{row}"].alignment = Alignment(horizontal="right")

    # Ancho columnas
    hoja_resumen.column_dimensions['A'].width = 30
    hoja_resumen.column_dimensions['B'].width = 20

    wb.save('uploads/reporte_financiero.xlsx')

    return redirect('/')

if __name__ == '__main__':
    app.run(debug=True)

